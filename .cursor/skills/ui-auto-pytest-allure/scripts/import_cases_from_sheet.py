#!/usr/bin/env python3
"""Import UI cases from Excel (.xlsx) or CSV (Feishu export) into NL specs and pytest tests."""

from __future__ import annotations

import argparse
import csv
import json
import subprocess
import sys
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))

from project_paths import resolve_project_root  # noqa: E402

PROJECT_ROOT = resolve_project_root(SCRIPT_DIR)

from generate_ui_test import validate_spec, write_outputs  # noqa: E402
from nl_case_parser import parse_natural_language_case  # noqa: E402
from sheet_case_parser import rows_to_nl_cases  # noqa: E402


def _read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.reader(handle)
                rows = list(reader)
            break
        except UnicodeDecodeError:
            rows = None
    else:
        raise SystemExit(f"Could not decode CSV file: {path}")

    if not rows:
        raise SystemExit(f"CSV file is empty: {path}")
    return rows[0], rows[1:]


def _read_xlsx(path: Path) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "Reading .xlsx requires openpyxl. Install with:\n"
            "  pip install openpyxl"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(["" if cell is None else str(cell).strip() for cell in row])
    workbook.close()

    if not rows:
        raise SystemExit(f"Excel file is empty: {path}")
    return rows[0], rows[1:]


def _load_sheet(path: Path) -> tuple[list[str], list[list[str]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    raise SystemExit(f"Unsupported file type: {suffix}. Use .csv or .xlsx")


def _write_nl_file(cases_root: Path, nl_text: str, test_name: str, priority: str) -> Path:
    out_path = cases_root / f"{test_name}.nl"
    out_path.write_text(nl_text, encoding="utf-8")
    return out_path


def _generate_from_nl(
    nl_text: str,
    spec_root: Path,
    output_root: Path,
    skip_install: bool,
) -> tuple[Path, Path]:
    spec = parse_natural_language_case(nl_text)
    validate_spec(spec)

    priority = spec["priority"]
    spec_path = spec_root / priority / f"{spec['test_name']}.json"
    spec_path.parent.mkdir(parents=True, exist_ok=True)
    spec_path.write_text(json.dumps(spec, ensure_ascii=False, indent=2), encoding="utf-8")

    output_dir = output_root / priority
    test_path = write_outputs(spec, output_dir)

    if not skip_install:
        installer = SCRIPT_DIR / "install_ui_dependencies.py"
        subprocess.run([sys.executable, str(installer)], check=False)

    return spec_path, test_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import cases from Excel/CSV (Feishu cloud sheet export supported)."
    )
    parser.add_argument(
        "sheet",
        type=Path,
        help="Path to .csv or .xlsx file exported from Excel / 飞书云表格",
    )
    parser.add_argument(
        "--cases-root",
        type=Path,
        default=PROJECT_ROOT / "cases",
        help="Directory to write .nl case files.",
    )
    parser.add_argument("--spec-root", type=Path, default=PROJECT_ROOT / "specs")
    parser.add_argument("--output-root", type=Path, default=PROJECT_ROOT / "generated-tests/ui")
    parser.add_argument(
        "--nl-only",
        action="store_true",
        help="Only write .nl files, do not generate specs and pytest scripts.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and print results without writing files.",
    )
    parser.add_argument("--skip-install", action="store_true")
    args = parser.parse_args()

    sheet_path = args.sheet.resolve()
    if not sheet_path.exists():
        raise SystemExit(f"Sheet file not found: {sheet_path}")

    headers, data_rows = _load_sheet(sheet_path)
    cases = rows_to_nl_cases(data_rows, headers)

    print(f"Loaded {len(cases)} case(s) from {sheet_path.name}")
    print("Required headers: 用例等级、用例名、功能模块、子模块、操作步骤")
    print()

    for row_number, test_name, nl_text in cases:
        print(f"--- Row {row_number}: {test_name} ---")
        if args.dry_run:
            print(nl_text)
            print()
            continue

        nl_path = _write_nl_file(args.cases_root, nl_text, test_name, "")
        print(f"  NL: {nl_path}")

        if args.nl_only:
            continue

        spec_path, test_path = _generate_from_nl(
            nl_text,
            args.spec_root,
            args.output_root,
            args.skip_install,
        )
        print(f"  Spec: {spec_path}")
        print(f"  Test: {test_path}")
        print()

    if args.dry_run:
        print("Dry run complete. No files written.")
    else:
        print(f"Imported {len(cases)} case(s).")


if __name__ == "__main__":
    main()
